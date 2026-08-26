import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    KeepTogether,
)
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend


class NumberedCanvas(canvas.Canvas):
    """Two-pass canvas to dynamically compute and render total page count."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_footer(num_pages)
            super().showPage()
        super().save()

    def draw_page_footer(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))

        # Footer divider line
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.75)
        self.line(36, 36, 792 - 36, 36)

        # Left footer: System branding & generation notice
        self.drawString(36, 24, "Eventify Analytics & Reporting — Confidential")

        # Right footer: Page number
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(792 - 36, 24, page_text)
        self.restoreState()


def _get_local_timestamp_str(dt=None):
    """Format datetime with local date and time (e.g. August 25, 2026 at 18:08)."""
    if dt is None:
        local_dt = datetime.now().astimezone()
    elif timezone.is_aware(dt):
        local_dt = dt.astimezone()
    else:
        local_dt = dt

    return local_dt.strftime("%B %d, %Y at %H:%M")


def export_events_excel(events, user=None):
    """
    Generate an Excel (.xlsx) workbook containing detailed event sales,
    capacity metrics, ticket status, and total revenue calculations.
    All data, header, and total rows are centered.
    Rows 1 and 2 have clean borders.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events Performance"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palettes
    primary_color = "4F46E5"     # Indigo 600
    primary_dark = "3730A3"      # Indigo 800
    accent_emerald = "059669"    # Emerald 600
    gray_light = "F8FAFC"        # Slate 50
    border_gray = "CBD5E1"       # Slate 300

    font_title = Font(name="Calibri", size=16, bold=True, color=primary_dark)
    font_subtitle = Font(name="Calibri", size=9, italic=True, color="64748B")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10, color="1E293B")
    font_revenue = Font(name="Calibri", size=10, bold=True, color=accent_emerald)
    font_totals = Font(name="Calibri", size=11, bold=True, color="0F172A")

    fill_header = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    fill_zebra = PatternFill(start_color=gray_light, end_color=gray_light, fill_type="solid")
    fill_totals = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=border_gray),
        right=Side(style='thin', color=border_gray),
        top=Side(style='thin', color=border_gray),
        bottom=Side(style='thin', color=border_gray),
    )
    total_top_border = Side(style='thick', color=primary_dark)
    total_bottom_border = Side(style='double', color=primary_dark)
    totals_border = Border(
        left=Side(style='thin', color=border_gray),
        right=Side(style='thin', color=border_gray),
        top=total_top_border,
        bottom=total_bottom_border,
    )

    # 1. Report Title Header (Centered across columns with distinct visible borders)
    title_cell = ws["A1"]
    title_cell.value = "Eventify — Event Sales & Performance Analytics"
    title_cell.font = font_title
    ws.row_dimensions[1].height = 28

    gen_time = _get_local_timestamp_str()
    account_info = user.username if user else "All Accounts"
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {gen_time}  |  Account: {account_info}  |  Total Events: {len(events)}"
    sub_cell.font = font_subtitle
    ws.row_dimensions[2].height = 18

    # Blank row
    ws.row_dimensions[3].height = 10

    # 2. Table Column Headers (Centered)
    headers = [
        "Event ID",
        "Event Title",
        "Event Date & Time",
        "Location",
        "Category",
        "Ticket Price ($)",
        "Max Capacity",
        "Tickets Sold",
        "Tickets Remaining",
        "Total Revenue ($)",
        "Status",
    ]

    # Normal black borders and centered alignment for rows 1 and 2
    black_thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000"),
    )
    num_headers = len(headers)

    for col_idx in range(1, num_headers + 1):
        c1 = ws.cell(row=1, column=col_idx)
        c1.alignment = Alignment(horizontal="centerContinuous", vertical="center")
        c1.border = black_thin_border

        c2 = ws.cell(row=2, column=col_idx)
        c2.alignment = Alignment(horizontal="centerContinuous", vertical="center")
        c2.border = black_thin_border

    start_row = 4
    ws.row_dimensions[start_row].height = 24

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header_title
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 3. Data Rows (All Centered)
    total_capacity = 0
    total_sold = 0
    total_remaining = 0
    total_revenue = Decimal("0.00")

    current_row = start_row + 1
    for index, event in enumerate(events):
        sold = event.tickets_sold
        remaining = event.tickets_remaining
        rev = event.revenue
        capacity = event.max_tickets

        total_capacity += capacity
        total_sold += sold
        total_remaining += remaining
        total_revenue += rev

        if event.is_sold_out:
            status_text = "Sold Out"
        elif event.is_expired:
            status_text = "Expired / Ended"
        elif remaining <= 5:
            status_text = f"Almost Full ({remaining} left)"
        else:
            status_text = "Active"

        if event.date:
            local_dt = event.date.astimezone() if timezone.is_aware(event.date) else event.date
            date_str = local_dt.strftime("%Y-%m-%d %H:%M")
        else:
            date_str = "N/A"
        category_name = event.category_label

        row_values = [
            f"#{event.pk}",
            event.title,
            date_str,
            event.location or "Online / TBD",
            category_name,
            float(event.price),
            capacity,
            sold,
            remaining,
            float(rev),
            status_text,
        ]

        ws.row_dimensions[current_row].height = 20
        is_even = (index % 2 == 1)

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = val
            cell.font = font_data
            cell.border = thin_border
            # ALL data cells are centered
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if is_even:
                cell.fill = fill_zebra

            # Number Formats
            if col_num == 6:  # Price
                cell.number_format = "$#,##0.00"
            elif col_num in (7, 8, 9):  # Capacity, Sold, Remaining
                cell.number_format = "#,##0"
            elif col_num == 10:  # Revenue
                cell.font = font_revenue
                cell.number_format = "$#,##0.00"

        current_row += 1

    # 4. Totals / Summary Row (All Centered)
    ws.row_dimensions[current_row].height = 24
    totals_row = [
        "TOTALS",
        f"{len(events)} Event(s)",
        "",
        "",
        "",
        "",
        total_capacity,
        total_sold,
        total_remaining,
        float(total_revenue),
        "",
    ]

    for col_num, val in enumerate(totals_row, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = val
        cell.font = font_totals
        cell.fill = fill_totals
        cell.border = totals_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if col_num in (7, 8, 9):
            cell.number_format = "#,##0"
        elif col_num == 10:
            cell.font = Font(name="Calibri", size=11, bold=True, color=accent_emerald)
            cell.number_format = "$#,##0.00"

    # 5. Column Width Auto-Fitting
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.row > 2:  # Skip merged title rows
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    # Save to memory buffer
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"eventify_analytics_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_events_pdf(events, total_events, total_tickets_sold, total_tickets_remaining, total_revenue, user=None):
    """
    Generate a styled PDF analytics report with revenue metrics,
    KPI summary cards, dual visual charts (Bar + Donut), and detailed event breakdown.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=48,
    )

    styles = getSampleStyleSheet()

    # Custom typography styles
    color_primary = colors.HexColor("#4F46E5")   # Indigo
    color_dark = colors.HexColor("#0F172A")      # Slate 900
    color_text = colors.HexColor("#334155")      # Slate 700
    color_muted = colors.HexColor("#64748B")     # Slate 500
    color_emerald = colors.HexColor("#059669")   # Emerald 600

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=color_primary,
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=color_muted,
    )
    section_heading_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=15,
        textColor=color_dark,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=color_text,
    )
    cell_bold_style = ParagraphStyle(
        'TableCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=color_dark,
    )
    cell_revenue_style = ParagraphStyle(
        'TableCellRevenue',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=color_emerald,
        alignment=1,  # Center
    )

    # KPI Typography Styles (Clean Centering)
    kpi_label_style = ParagraphStyle(
        'KPILabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=color_muted,
        alignment=1,  # Center
    )
    kpi_val_dark_style = ParagraphStyle(
        'KPIValDark',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=color_dark,
        alignment=1,  # Center
    )
    kpi_val_sold_style = ParagraphStyle(
        'KPIValSold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=color_primary,
        alignment=1,  # Center
    )
    kpi_val_rev_style = ParagraphStyle(
        'KPIValRev',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=color_emerald,
        alignment=1,  # Center
    )

    story = []

    # 1. Header Banner
    account_name = user.username if user else "All Accounts"
    gen_time_str = _get_local_timestamp_str()

    header_table_data = [
        [
            Paragraph("Eventify — Analytics & Performance Report", title_style),
            Paragraph(f"<b>Generated:</b> {gen_time_str}<br/><b>Account:</b> {account_name}", subtitle_style),
        ]
    ]
    header_table = Table(header_table_data, colWidths=[480, 240])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(header_table)

    # Accent Divider line
    d_line = Drawing(720, 4)
    d_line.add(Line(0, 2, 720, 2, strokeColor=color_primary, strokeWidth=2))
    story.append(d_line)
    story.append(Spacer(1, 10))

    # 2. Key Metrics Summary (KPI Cards - Perfectly Centered & Positioned)
    kpi_card_data = [
        [
            [Paragraph("TOTAL EVENTS", kpi_label_style), Spacer(1, 3), Paragraph(str(total_events), kpi_val_dark_style)],
            [Paragraph("TICKETS SOLD", kpi_label_style), Spacer(1, 3), Paragraph(str(total_tickets_sold), kpi_val_sold_style)],
            [Paragraph("TICKETS REMAINING", kpi_label_style), Spacer(1, 3), Paragraph(str(total_tickets_remaining), kpi_val_dark_style)],
            [Paragraph("TOTAL REVENUE", kpi_label_style), Spacer(1, 3), Paragraph(f"${float(total_revenue):,.2f}", kpi_val_rev_style)],
        ]
    ]
    kpi_table = Table(kpi_card_data, colWidths=[172.5, 172.5, 172.5, 172.5])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor("#F8FAFC")),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor("#EEF2FF")),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor("#F8FAFC")),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor("#ECFDF5")),
        ('BOX', (0, 0), (0, 0), 1, colors.HexColor("#CBD5E1")),
        ('BOX', (1, 0), (1, 0), 1, colors.HexColor("#C7D2FE")),
        ('BOX', (2, 0), (2, 0), 1, colors.HexColor("#CBD5E1")),
        ('BOX', (3, 0), (3, 0), 1, colors.HexColor("#A7F3D0")),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))

    # 3. Dual Visual Analytics Charts (Side-by-Side: Bar Chart + Donut Chart)
    if events and len(events) > 0:
        story.append(Paragraph("Performance Breakdown & Sales Volume", section_heading_style))

        # Chart A: Column Bar Chart (Smaller, with explicit X-axis & Y-axis labels)
        chart_events = list(reversed(list(events)))[:6]
        labels = [e.title[:12] + ("…" if len(e.title) > 12 else "") for e in chart_events]
        sold_data = [e.tickets_sold for e in chart_events]
        capacity_data = [e.max_tickets for e in chart_events]

        d_bar = Drawing(430, 135)
        d_bar.add(Rect(0, 0, 430, 135, fillColor=colors.HexColor("#F8FAFC"), strokeColor=colors.HexColor("#E2E8F0"), strokeWidth=1, rx=6, ry=6))
        d_bar.add(String(12, 120, "Tickets Sold vs Capacity", fontName='Helvetica-Bold', fontSize=8, fillColor=colors.HexColor("#0F172A")))

        # Y-Axis Label (Lighter purple, top of Y axis - moved left)
        d_bar.add(String(6, 100, "Tickets", fontName='Helvetica-Bold', fontSize=7, fillColor=colors.HexColor("#6366F1")))

        # X-Axis Label (Lighter purple, right end of X axis)
        d_bar.add(String(316, 14, "Events", fontName='Helvetica-Bold', fontSize=7, fillColor=colors.HexColor("#6366F1")))

        bc = VerticalBarChart()
        bc.x = 46
        bc.y = 18
        bc.height = 80
        bc.width = 266
        bc.data = [sold_data, capacity_data]
        bc.categoryAxis.categoryNames = labels
        bc.categoryAxis.labels.fontSize = 7
        bc.categoryAxis.labels.dy = -8
        bc.categoryAxis.labels.fontName = 'Helvetica'
        bc.categoryAxis.visibleGrid = 0

        bc.valueAxis.valueMin = 0
        max_val = max(capacity_data + [1])
        bc.valueAxis.valueMax = max(max_val + (2 if max_val < 10 else int(max_val * 0.2)), 5)
        bc.valueAxis.valueStep = max(int(bc.valueAxis.valueMax / 4), 1)
        bc.valueAxis.labels.fontSize = 7
        bc.valueAxis.labels.fontName = 'Helvetica'
        bc.valueAxis.visibleGrid = 1
        bc.valueAxis.gridStrokeColor = colors.HexColor("#E2E8F0")

        bc.bars[0].fillColor = colors.HexColor("#4F46E5")  # Tickets Sold
        bc.bars[1].fillColor = colors.HexColor("#CBD5E1")  # Capacity
        d_bar.add(bc)

        # Bar Chart Legend
        bar_legend = Legend()
        bar_legend.fontName = 'Helvetica'
        bar_legend.fontSize = 7.5
        bar_legend.x = 328
        bar_legend.y = 82
        bar_legend.dx = 8
        bar_legend.dy = 8
        bar_legend.dxTextSpace = 5
        bar_legend.yGap = 5
        bar_legend.colorNamePairs = [
            (colors.HexColor("#4F46E5"), 'Tickets Sold'),
            (colors.HexColor("#CBD5E1"), 'Total Capacity'),
        ]
        d_bar.add(bar_legend)

        # Chart B: Donut Chart (Ticket Capacity Allocation: Sold vs Remaining)
        d_donut = Drawing(275, 135)
        d_donut.add(Rect(0, 0, 275, 135, fillColor=colors.HexColor("#F8FAFC"), strokeColor=colors.HexColor("#E2E8F0"), strokeWidth=1, rx=6, ry=6))
        d_donut.add(String(12, 120, "Ticket Capacity Allocation", fontName='Helvetica-Bold', fontSize=8, fillColor=colors.HexColor("#0F172A")))

        total_cap = total_tickets_sold + total_tickets_remaining
        sold_pct = int((total_tickets_sold / total_cap * 100)) if total_cap > 0 else 0

        pie = Pie()
        pie.x = 18
        pie.y = 14
        pie.width = 80
        pie.height = 80
        pie.data = [total_tickets_sold, total_tickets_remaining] if total_cap > 0 else [0, 1]
        pie.slices[0].fillColor = colors.HexColor("#4F46E5")  # Sold
        pie.slices[1].fillColor = colors.HexColor("#CBD5E1")  # Remaining
        pie.slices.strokeColor = colors.white
        pie.slices.strokeWidth = 1.5
        d_donut.add(pie)

        # Donut center cutout circle & percentage
        d_donut.add(Circle(58, 54, 23, fillColor=colors.HexColor("#F8FAFC"), strokeColor=colors.white, strokeWidth=1))
        pct_text = f"{sold_pct}%"
        d_donut.add(String(50 if len(pct_text) > 3 else 52, 51, pct_text, fontName='Helvetica-Bold', fontSize=7.5, fillColor=colors.HexColor("#4F46E5")))

        # Donut Legend
        pie_legend = Legend()
        pie_legend.fontName = 'Helvetica'
        pie_legend.fontSize = 7.5
        pie_legend.x = 115
        pie_legend.y = 72
        pie_legend.dx = 8
        pie_legend.dy = 8
        pie_legend.dxTextSpace = 5
        pie_legend.yGap = 5
        pie_legend.colorNamePairs = [
            (colors.HexColor("#4F46E5"), f"Sold: {total_tickets_sold}"),
            (colors.HexColor("#CBD5E1"), f"Remaining: {total_tickets_remaining}"),
        ]
        d_donut.add(pie_legend)
        d_donut.add(String(115, 26, f"Total Capacity: {total_cap}", fontName='Helvetica', fontSize=7.5, fillColor=colors.HexColor("#64748B")))

        # Place charts side by side in a table
        charts_table = Table([[d_bar, d_donut]], colWidths=[438, 282])
        charts_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(charts_table)
        story.append(Spacer(1, 10))

    # 4. Detailed Event Sales Table (All Cleanly Centered / Aligned)
    story.append(Paragraph("Event Sales & Revenue Breakdown", section_heading_style))

    table_headers = [
        Paragraph("<b>Event Title</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Date & Time</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Category</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Price</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Sold / Cap</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Remaining</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Total Revenue</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
        Paragraph("<b>Status</b>", ParagraphStyle('TH', parent=cell_bold_style, textColor=colors.white, alignment=1)),
    ]

    table_data = [table_headers]

    for event in events:
        sold = event.tickets_sold
        remaining = event.tickets_remaining
        rev = event.revenue

        if event.is_sold_out:
            status_html = "<font color='#DC2626'><b>Sold Out</b></font>"
        elif event.is_expired:
            status_html = "<font color='#64748B'>Expired</font>"
        elif remaining <= 5:
            status_html = f"<font color='#D97706'><b>{remaining} left</b></font>"
        else:
            status_html = "<font color='#059669'>Active</font>"

        if event.date:
            local_dt = event.date.astimezone() if timezone.is_aware(event.date) else event.date
            date_formatted = local_dt.strftime("%b %d, %Y @ %H:%M")
        else:
            date_formatted = "N/A"

        row = [
            Paragraph(f"<b>{event.title}</b><br/><font size='7' color='#64748B'>{event.location or 'Online'}</font>", ParagraphStyle('TC_Title', parent=cell_style, alignment=1)),
            Paragraph(date_formatted, ParagraphStyle('TC_C', parent=cell_style, alignment=1)),
            Paragraph(event.category_label, ParagraphStyle('TC_C', parent=cell_style, alignment=1)),
            Paragraph(f"${event.price}", ParagraphStyle('TC_C', parent=cell_style, alignment=1)),
            Paragraph(f"<b>{sold}</b> / {event.max_tickets}", ParagraphStyle('TC_C', parent=cell_style, alignment=1)),
            Paragraph(str(remaining), ParagraphStyle('TC_C', parent=cell_style, alignment=1)),
            Paragraph(f"<b>${float(rev):,.2f}</b>", cell_revenue_style),
            Paragraph(status_html, ParagraphStyle('TC_C', parent=cell_style, alignment=1)),
        ]
        table_data.append(row)

    # Summary Totals Row in PDF Table
    totals_row = [
        Paragraph("<b>TOTALS</b>", ParagraphStyle('TC_TotTitle', parent=cell_bold_style, alignment=1)),
        Paragraph(f"<b>{len(events)} Events</b>", ParagraphStyle('TC_Tot', parent=cell_bold_style, alignment=1)),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph(f"<b>{total_tickets_sold}</b>", ParagraphStyle('TC_Tot', parent=cell_bold_style, alignment=1)),
        Paragraph(f"<b>{total_tickets_remaining}</b>", ParagraphStyle('TC_Tot', parent=cell_bold_style, alignment=1)),
        Paragraph(f"<b>${float(total_revenue):,.2f}</b>", cell_revenue_style),
        Paragraph("", cell_style),
    ]
    table_data.append(totals_row)

    # Column widths total 720pt (landscape letter width 792 - 72 margins)
    col_widths = [180, 105, 80, 55, 75, 65, 85, 75]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), color_primary),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('INNERGRID', (0, 0), (-1, -2), 0.5, colors.HexColor("#F1F5F9")),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
    ]

    # Alternate row colors
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            t_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F8FAFC")))

    # Summary row styling
    t_style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#EEF2FF")))
    t_style.append(('LINEABOVE', (0, -1), (-1, -1), 1.5, color_primary))
    t_style.append(('BOTTOMPADDING', (0, -1), (-1, -1), 6))
    t_style.append(('TOPPADDING', (0, -1), (-1, -1), 6))

    table.setStyle(TableStyle(t_style))
    story.append(table)

    # Build PDF with NumberedCanvas
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)

    filename = f"eventify_analytics_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
